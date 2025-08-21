import os

import openpyxl
from openpyxl.styles import PatternFill

from src.data_collection_bot.config import DATA_TABLE_PATH


async def save_records(
        user_pseudonym: str,
        answers: dict
):
    date_set: set[str] = {rec['date'] for rec in answers.values()}
    if len(date_set) != 1:
        raise ValueError("More than one date")

    date: str = list(date_set)[0]

    parameters: list[str] = list(answers.keys())
    param_to_answer = {k: v['answer'] for k, v in answers.items()}
    param_to_norm = {k: v['is_norm'] for k, v in answers.items()}

    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    if os.path.isfile(DATA_TABLE_PATH):
        book = openpyxl.load_workbook(DATA_TABLE_PATH)
        if 'init' in book.sheetnames:
            init_sheet = book['init']
            book.remove(init_sheet)
        if user_pseudonym in book.sheetnames:
            sheet = book[user_pseudonym]
        else:
            sheet = book.create_sheet(user_pseudonym)
    else:
        book = openpyxl.Workbook()
        book.remove(book.active)
        sheet = book.create_sheet(user_pseudonym)

    if sheet.max_row == 1:
        sheet.cell(row=1, column=1).value = 'Параметр'
        for i, param in enumerate(parameters):
            sheet.cell(row=i+2, column=1).value = param

    date_col = None
    for col in  range(2, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == date:
            date_col = col
            break

    if date_col is None:
        date_col = sheet.max_column + 1
        sheet.cell(row=1, column=date_col).value = date

    param_row_name = {}
    for row in range(2, sheet.max_row + 1):
        param_name = sheet.cell(row=row, column=1).value
        param_row_name[param_name] = row

    for param in parameters:
        row = param_row_name[param]
        sheet.cell(row=row, column=date_col).value = param_to_answer[param]
        if not param_to_norm[param]:
            sheet.cell(row=row, column=date_col).fill = fill

    book.save(DATA_TABLE_PATH)
